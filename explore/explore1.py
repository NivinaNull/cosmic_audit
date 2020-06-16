#-*- coding: utf-8 -*-
# @Time    : 2020/6/5 15:48
# @Author  : Nivina
# @Email   : nivinanull@163.com
# @File    : explore1.py


from openpyxl import load_workbook, Workbook

wb001 = load_workbook("./data/附件7：需求序号004：COSMIC软件评估功能点拆分表.xlsx")
# print(wb001.get_sheet_names())
sheet_name = ''
wb = Workbook()
ws = wb.create_sheet()
for i in wb001.get_sheet_names():
    if  '功能点' in i:sheet_name = i
if sheet_name:
    # print(sheet_name)
    ws = wb001.get_sheet_by_name(sheet_name)
else:
    print('无功能点拆分表')

column = 0
for i in ws[1]:
    if i.value == '子过程描述':
        column = i.column
        # print(column)
# for col in ws.iter_cols(min_row=1, min_col=column, max_col=column, max_row=10):
#     for cell in col:
#         print(cell)

col = chr(65 + column - 1)
# print(col)
detail_list = []
for cell in ws[col]:
    detail_list.append(cell.value)
del detail_list[0]

print(detail_list)
print(len(detail_list))
print(len(set(detail_list)))

# col_end = 'A'+str(len(ws.columns))
# cols = ws['A1':col_end]
# print(cols)

# for row in ws.iter_rows(min_row=1,max_row=1, values_only=True):
#     print(row)
#     print(type(row))



